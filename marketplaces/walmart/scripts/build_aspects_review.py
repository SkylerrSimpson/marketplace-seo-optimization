#!/usr/bin/env python3
"""
build_aspects_review.py -- the ONE aspects review sheet, matching eBay's
item-specifics-aspects_REVIEW.csv pattern: every applicable aspect for every SKU in
a single file, filled or blank, so a reviewer can go down it row by row and fill in
whatever's missing. Replaces the earlier two-file split (aspects_us.csv with only
filled values + a separate missing_aspects_us.csv gap summary) -- that split wasn't
what was wanted; this is the unified version.

For each SKU:
  1. Look up its product type's full applicable-attribute set (name + human title)
     from product_type_specs_{country}.json (built by fetch_product_type_specs.php --
     the real Get Spec API response per product type, not the mixed-batch xlsx, which
     only lists attributes relevant to whichever up-to-10 product types happened to
     share that particular export batch).
  2. Look up whatever values that SKU's own xlsx batch file actually has for each of
     those attributes (blank if the attribute wasn't filled in, or wasn't even
     exported as a column in that file -- both cases mean "needs review").
  3. Look up the Required/Recommended label for that attribute FROM THAT SAME BATCH
     FILE's hidden Requirement Level table, when available.

One row per (sku, attribute, value) -- multi-value attributes (e.g. "Accessories
Included") get one row per filled value; an attribute with zero values gets exactly
one row with a blank current_value.

Usage: python3 marketplaces/walmart/scripts/build_aspects_review.py --country=us
Output: marketplaces/walmart/data/{country}/output/aspects_review_{country}.csv
  sku, wpid, title, varied_by, aspect_name, requirement_level,
  current_value, proposed_value, approved_value, reviewer_notes
"""

import argparse
import csv
import glob
import json
import os

import openpyxl

from merge_aspects_from_bulk_export import (
    REPO_ROOT,
    TARGET_SECTION,
    resolve_columns,
    build_aspect_plan,
)
from find_missing_aspects import read_requirement_levels


def collect_real_aspect_xml_names(files):
    """
    Get Spec's schema lists EVERY 'Visible' property for a product type, including
    structural/admin fields (variantAttributeNames, variantGroupId, mainImageUrl,
    price, etc.) that live outside the "Product Content to improve search & browse"
    section in the xlsx export and were never meant to be reviewed as if they were
    product specs. Since Get Spec's schema doesn't carry that section boundary, build
    the allowlist the other way: scan every xlsx header (fast -- headers only, no data
    rows) and collect every req_xml that ever appeared under TARGET_SECTION across all
    batches. Anything not in this set gets excluded from the review sheet even if
    Get Spec lists it as applicable to a product type.
    """
    allowed = set()
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if "Product Content And Site Exp" not in wb.sheetnames:
            continue
        cols = resolve_columns(wb["Product Content And Site Exp"])
        for emitter in build_aspect_plan(cols):
            if emitter.get("req_xml"):
                allowed.add(emitter["req_xml"])
        wb.close()
    return allowed


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--country", default="us")
    args = ap.parse_args()
    country = args.country.lower()

    input_dir = os.path.join(REPO_ROOT, "walmart", "data", country, "input")
    output_dir = os.path.join(REPO_ROOT, "walmart", "data", country, "output")
    listings_path = os.path.join(input_dir, "listings.json")
    specs_path = os.path.join(output_dir, f"product_type_specs_{country}.json")

    if not os.path.isfile(specs_path):
        raise SystemExit(
            f"missing {specs_path} -- run fetch_product_type_specs.php --country={country} first"
        )
    with open(specs_path) as fh:
        specs_by_type = json.load(fh)  # {productType: {xml_name: title}}

    with open(listings_path) as fh:
        listings = json.load(fh)
    wpid_of = {l["sku"]: l.get("wpid", "") for l in listings}
    title_of = {l["sku"]: l.get("title", "") for l in listings}
    product_type_of = {l["sku"]: l.get("productType", "") for l in listings}
    all_skus = set(wpid_of.keys())

    # --- pass 1: per-SKU export data (current values + requirement levels), keyed
    # off whichever xlsx batch file that SKU actually appears in.
    export_values = {}   # sku -> {req_xml: [values...]}
    export_varied = {}   # sku -> "Color; Size"
    req_levels_of_sku = {}  # sku -> {req_xml: 'Required'|'Recommended'}
    covered_skus = set()

    files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    print(f"found {len(files)} xlsx files in {input_dir}")

    real_aspect_xml_names = collect_real_aspect_xml_names(files)
    print(f"real product-content attributes (section={TARGET_SECTION!r}): {len(real_aspect_xml_names)}")

    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if "Product Content And Site Exp" not in wb.sheetnames:
            continue
        ws = wb["Product Content And Site Exp"]
        cols = resolve_columns(ws)
        plan = build_aspect_plan(cols)
        req_levels = read_requirement_levels(wb)
        variant_cols = [c["idx"] for c in cols if c["xml"] == "variantAttributeNames"]
        sku_idx = next(c["idx"] for c in cols if c["xml"] == "sku")

        for row in ws.iter_rows(min_row=7, values_only=True):
            sku = row[sku_idx]
            if not sku or sku in covered_skus:
                continue
            covered_skus.add(sku)
            req_levels_of_sku[sku] = req_levels

            varying = {str(row[i]).strip() for i in variant_cols if row[i]}
            varied_names = []
            values = {}
            for emitter in plan:
                req_xml = emitter.get("req_xml")
                if req_xml is None:
                    continue
                if emitter["kind"] == "simple":
                    val = row[emitter["idx"]]
                    is_var = req_xml in varying
                else:
                    mval, uval = row[emitter["measure_idx"]], row[emitter["unit_idx"]]
                    val = " ".join(str(x) for x in (mval, uval) if x not in (None, "")) or None
                    is_var = False
                if val not in (None, ""):
                    values.setdefault(req_xml, []).append(str(val))
                    if is_var and emitter["name"] not in varied_names:
                        varied_names.append(emitter["name"])
            export_values[sku] = values
            export_varied[sku] = "; ".join(varied_names)
        wb.close()

    print(f"export data found for {len(covered_skus)}/{len(all_skus)} SKUs "
          f"({len(all_skus) - len(covered_skus)} have no batch file -- current_value/"
          f"requirement_level will be blank for those)")

    # --- pass 2: emit one row per SKU x applicable attribute ------------------------
    out_rows = 0
    out_path = os.path.join(output_dir, f"aspects_review_{country}.csv")
    with open(out_path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow([
            "sku", "wpid", "title", "varied_by", "aspect_name", "requirement_level",
            "current_value", "proposed_value", "approved_value", "reviewer_notes",
        ])
        for sku in sorted(all_skus):
            ptype = product_type_of.get(sku, "")
            applicable = specs_by_type.get(ptype, {})
            if not applicable:
                continue  # product type wasn't in the spec cache (shouldn't happen if fetch covered all types)

            values = export_values.get(sku, {})
            req_levels = req_levels_of_sku.get(sku, {})
            varied_by = export_varied.get(sku, "")

            for xml_name, human_title in applicable.items():
                if xml_name not in real_aspect_xml_names:
                    continue  # structural/admin field (variantAttributeNames, mainImageUrl, price, ...), not a reviewable spec
                level = req_levels.get(xml_name, "")
                vals = values.get(xml_name, [])
                if vals:
                    for v in vals:
                        w.writerow([sku, wpid_of.get(sku, ""), title_of.get(sku, ""),
                                    varied_by, human_title, level, v, "", "", ""])
                        out_rows += 1
                else:
                    w.writerow([sku, wpid_of.get(sku, ""), title_of.get(sku, ""),
                                varied_by, human_title, level, "", "", "", ""])
                    out_rows += 1

    print(f"\ndone: {len(all_skus)} SKUs, {out_rows} aspect rows -> {out_path}")


if __name__ == "__main__":
    main()

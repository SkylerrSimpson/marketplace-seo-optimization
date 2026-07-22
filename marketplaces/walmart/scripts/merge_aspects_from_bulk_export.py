#!/usr/bin/env python3
"""
merge_aspects_from_bulk_export.py -- merge Walmart Seller Center's bulk "Update with
GTINs" export (.xlsx files dropped in walmart/data/{country}/input/) into
aspects_{country}.csv.

Why this exists: Walmart's read APIs (getAllItems, Item Search, the Item Report) do
NOT expose the full attribute/spec set (Color, Material, Assembled Product Weight,
etc.) -- only the Seller Center bulk "Update with GTINs" export comes back
pre-populated with current values. See walmart/README.md, "Aspects/attributes have NO
read-back endpoint" -- confirmed against 10,222 real items across four API surfaces
before falling back to this export.

Each downloaded .xlsx has a DIFFERENT column layout per file -- Walmart only includes
attribute columns relevant to the product types in that batch, so column positions
cannot be hardcoded. Every file's header is resolved independently:
  row 2: section label (merged cell, forward-filled across blanks)
  row 3: attribute-group label, only populated for Measure/Unit pairs (also
         forward-filled -- e.g. col N = "Assembled Product Weight (assembledProductWeight)"
         then col N+1 blank belongs to the same group)
  row 4: human-readable column label (e.g. "Accessories Included 3 (+)")
  row 5: machine/xml attribute name (e.g. "accessoriesIncluded")
  row 7+: data

Only columns whose section is "Product Content to improve search & browse on Walmart
website" are treated as aspects -- this matches exactly what a human reviewer sees
under that heading in Seller Center; identifiers/images/pricing/logistics/compliance
columns are excluded.

Multi-value attributes (e.g. "Accessories Included" has 14 repeated columns) collapse
to one row per non-blank value, all sharing the same (stripped) aspect_name.
Measure/Unit pairs (e.g. "Assembled Product Weight") combine into one value like
"4.5 lb".

is_variant is set by cross-referencing each row's own variantAttributeNames column(s)
against that attribute's own xml name. NOTE: this uses the spec-template's naming
("color", "size") which does NOT match the read API's "actual_color"-style naming
used elsewhere in this pipeline (fetch_all_items.php) -- confirmed by inspecting real
export data; the two Walmart systems name the same concept differently.

Usage: python3 walmart/scripts/merge_aspects_from_bulk_export.py --country=us
Output: overwrites walmart/data/{country}/output/aspects_{country}.csv
"""

import argparse
import csv
import glob
import json
import os
import re

import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TARGET_SECTION = "Product Content to improve search & browse on Walmart website"


def forward_fill(row):
    out = []
    last = None
    for v in row:
        if v is not None:
            last = v
        out.append(last)
    return out


def clean_label(human):
    """'Accessories Included 3 (+)' -> 'Accessories Included'; 'Color' -> 'Color'."""
    s = re.sub(r"\s*\(\+\)\s*$", "", human)
    s = re.sub(r"\s+\d+$", "", s)
    return s.strip()


def resolve_columns(ws):
    """Read the 4 header rows and return a list of per-column metadata dicts."""
    rows = list(ws.iter_rows(min_row=2, max_row=5, values_only=True))
    section = forward_fill(rows[0])
    group = forward_fill(rows[1])
    human = rows[2]
    xml = rows[3]

    cols = []
    for i in range(len(xml)):
        cols.append({
            "idx": i,
            "section": section[i],
            "group": group[i],
            "human": human[i],
            "xml": xml[i],
        })
    return cols


def build_aspect_plan(cols):
    """
    Group the columns that fall in TARGET_SECTION into an ordered list of
    "aspect emitters": either a simple {kind:'simple', idx, name} or a
    {kind:'measure_unit', measure_idx, unit_idx, name} pair. Multi-value repeated
    columns (same cleaned label) all become separate 'simple' emitters sharing the
    same name -- that's exactly the desired one-row-per-value behavior already.
    """
    plan = []
    i = 0
    n = len(cols)
    while i < n:
        c = cols[i]
        if c["section"] != TARGET_SECTION or c["xml"] is None:
            i += 1
            continue
        if c["human"] == "Measure" and i + 1 < n and cols[i + 1]["human"] == "Unit":
            name = re.sub(r"\s*\([a-zA-Z_]+\)\s*(\(\+\))?\s*$", "", c["group"] or "measure").strip()
            req_xml = extract_group_xml(c["group"])
            plan.append({"kind": "measure_unit", "measure_idx": i, "unit_idx": i + 1, "name": name, "req_xml": req_xml})
            i += 2
            continue
        if c["human"] == "Unit" and i + 1 < n and cols[i + 1]["human"] == "Measure":
            # Net Content is the one field that reverses Unit/Measure order.
            name = re.sub(r"\s*\([a-zA-Z_]+\)\s*(\(\+\))?\s*$", "", c["group"] or "unit").strip()
            req_xml = extract_group_xml(c["group"])
            plan.append({"kind": "measure_unit", "measure_idx": i + 1, "unit_idx": i, "name": name, "req_xml": req_xml})
            i += 2
            continue
        if c["human"]:
            plan.append({"kind": "simple", "idx": i, "name": clean_label(c["human"]), "xml": c["xml"], "req_xml": c["xml"]})
        i += 1
    return plan


def extract_group_xml(group_label):
    """'Assembled Product Weight (assembledProductWeight)' -> 'assembledProductWeight'."""
    if not group_label:
        return None
    m = re.search(r"\(([a-zA-Z_]+)\)\s*(\(\+\))?\s*$", group_label)
    return m.group(1) if m else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--country", default="us")
    args = ap.parse_args()
    country = args.country.lower()

    input_dir = os.path.join(REPO_ROOT, "walmart", "data", country, "input")
    output_dir = os.path.join(REPO_ROOT, "walmart", "data", country, "output")
    listings_path = os.path.join(input_dir, "listings.json")

    with open(listings_path) as fh:
        listings = json.load(fh)
    wpid_of = {l["sku"]: l.get("wpid", "") for l in listings}

    files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    print(f"found {len(files)} xlsx files in {input_dir}")

    out_rows = []
    seen_skus = set()
    dup_skus = 0
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if "Product Content And Site Exp" not in wb.sheetnames:
            print(f"  SKIP (no data sheet): {os.path.basename(f)}")
            continue
        ws = wb["Product Content And Site Exp"]
        cols = resolve_columns(ws)
        plan = build_aspect_plan(cols)
        variant_cols = [c["idx"] for c in cols if c["xml"] == "variantAttributeNames"]
        sku_idx = next(c["idx"] for c in cols if c["xml"] == "sku")

        file_rows = 0
        for row in ws.iter_rows(min_row=7, values_only=True):
            sku = row[sku_idx]
            if not sku:
                continue
            if sku in seen_skus:
                dup_skus += 1
                continue  # a SKU appearing in >1 batch: first file wins, same data either way
            seen_skus.add(sku)

            varying = {str(row[i]).strip() for i in variant_cols if row[i]}
            wpid = wpid_of.get(sku, "")

            for emitter in plan:
                if emitter["kind"] == "simple":
                    val = row[emitter["idx"]]
                    if val in (None, ""):
                        continue
                    is_variant = "yes" if emitter["xml"] in varying else "no"
                    out_rows.append([sku, wpid, emitter["name"], str(val), is_variant])
                else:
                    mval = row[emitter["measure_idx"]]
                    uval = row[emitter["unit_idx"]]
                    if mval in (None, "") and uval in (None, ""):
                        continue
                    combined = " ".join(str(x) for x in (mval, uval) if x not in (None, ""))
                    out_rows.append([sku, wpid, emitter["name"], combined, "no"])
            file_rows += 1
        print(f"  {os.path.basename(f)}: {file_rows} SKUs, {len(plan)} aspect columns in this batch")
        wb.close()

    out_path = os.path.join(output_dir, f"aspects_{country}.csv")
    with open(out_path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["sku", "wpid", "aspect_name", "aspect_value", "is_variant"])
        w.writerows(out_rows)

    print(f"\ndone: {len(seen_skus)} SKUs, {dup_skus} duplicate-SKU rows skipped, {len(out_rows)} aspect rows -> {out_path}")


if __name__ == "__main__":
    main()

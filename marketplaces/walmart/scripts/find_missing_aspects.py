#!/usr/bin/env python3
"""
find_missing_aspects.py -- for every SKU, flag which attributes Walmart marks
Required (and separately, Recommended) for that item's batch but which are blank in
the Seller Center export.

Each downloaded .xlsx carries a hidden metadata sheet ("Hidden_product_content_and_sit")
with a per-attribute "Requirement Level" (Required | Recommended). That sheet's own
columns are laid out differently from the data sheet ("Product Content And Site Exp"),
so attributes are matched by their machine/xml name (e.g. "assembledProductWeight"),
not by column position -- see merge_aspects_from_bulk_export.py's build_aspect_plan()/
req_xml for how the xml name is recovered for Measure/Unit pairs.

IMPORTANT SCOPING FIX (2026-07-11): the hidden sheet's Requirement Level is scoped to
the whole BATCH (up to 10 mixed product types per export), not per item -- an
attribute required for one type in a batch (e.g. "Keychain Type") was showing up as
"missing required" on totally unrelated items (a gold panning kit, a paracord rope) in
the same batch. Confirmed against real data before shipping this. Fix: also require
the attribute to appear in that SKU's own product type's schema, fetched separately
per-type (not per-batch) via fetch_product_type_specs.php -> product_type_specs_
{country}.json. An attribute only counts as missing if it's BOTH labeled Required/
Recommended (from the xlsx) AND actually applicable to that item's specific product
type (from the per-type spec) -- run fetch_product_type_specs.php first.

Usage: python3 marketplaces/walmart/scripts/find_missing_aspects.py --country=us
Output: marketplaces/walmart/data/{country}/output/missing_aspects_{country}.csv
  sku, wpid, title, missing_required, missing_recommended
  (missing_required/missing_recommended are "; "-joined attribute names, blank if none)
"""

import argparse
import csv
import glob
import json
import os

import openpyxl

from merge_aspects_from_bulk_export import (
    REPO_ROOT,
    resolve_columns,
    build_aspect_plan,
)


def read_requirement_levels(wb):
    """xml_name -> 'Required' | 'Recommended', from the hidden metadata sheet."""
    if "Hidden_product_content_and_sit" not in wb.sheetnames:
        return {}
    ws = wb["Hidden_product_content_and_sit"]
    rows = list(ws.iter_rows(min_row=1, max_row=18, values_only=True))
    labels = [r[0] for r in rows]
    req_row = rows[labels.index("Requirement Level")]
    xml_row = rows[labels.index("Attribute XML Name")]
    levels = {}
    for xml_name, level in zip(xml_row[1:], req_row[1:]):
        if xml_name and level:
            levels[xml_name] = level
    return levels


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--country", default="us")
    args = ap.parse_args()
    country = args.country.lower()

    input_dir = os.path.join(REPO_ROOT, "walmart", "data", country, "input")
    output_dir = os.path.join(REPO_ROOT, "walmart", "data", country, "output")
    listings_path = os.path.join(input_dir, "listings.json")
    specs_path = os.path.join(output_dir, f"product_type_specs_{country}.json")

    if not os.path.isfile(specs_path):
        raise SystemExit(
            f"missing {specs_path} -- run fetch_product_type_specs.php --country={country} first "
            "(without it, missing-required results are unreliable across mixed-product-type batches)"
        )
    with open(specs_path) as fh:
        specs_by_type = json.load(fh)  # {productType: [xml_name, ...]}
    applicable_by_type = {t: set(attrs) for t, attrs in specs_by_type.items()}

    with open(listings_path) as fh:
        listings = json.load(fh)
    wpid_of = {l["sku"]: l.get("wpid", "") for l in listings}
    title_of = {l["sku"]: l.get("title", "") for l in listings}
    product_type_of = {l["sku"]: l.get("productType", "") for l in listings}

    files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    print(f"found {len(files)} xlsx files in {input_dir}")

    out_rows = []
    seen_skus = set()
    total_missing_required = 0
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if "Product Content And Site Exp" not in wb.sheetnames:
            continue
        ws = wb["Product Content And Site Exp"]
        cols = resolve_columns(ws)
        plan = build_aspect_plan(cols)
        req_levels = read_requirement_levels(wb)
        sku_idx = next(c["idx"] for c in cols if c["xml"] == "sku")

        for row in ws.iter_rows(min_row=7, values_only=True):
            sku = row[sku_idx]
            if not sku or sku in seen_skus:
                continue
            seen_skus.add(sku)

            # Collapse the plan (which has one emitter per repeated multi-value
            # column) down to one entry per underlying attribute for the blank-check
            # -- a multi-value attribute only counts as "missing" if EVERY one of its
            # repeated columns is blank, not just the first.
            filled = {}
            for emitter in plan:
                req_xml = emitter.get("req_xml")
                if req_xml is None:
                    continue
                if emitter["kind"] == "simple":
                    val = row[emitter["idx"]]
                else:
                    val = row[emitter["measure_idx"]]
                if req_xml not in filled:
                    filled[req_xml] = {"name": emitter["name"], "any_value": False}
                if val not in (None, ""):
                    filled[req_xml]["any_value"] = True

            applicable = applicable_by_type.get(product_type_of.get(sku, ""), set())

            missing_required = []
            missing_recommended = []
            for req_xml, level in req_levels.items():
                if req_xml not in applicable:
                    continue  # not applicable to THIS item's own product type
                info = filled.get(req_xml)
                if info is None:
                    continue  # this attribute column isn't even in this batch's plan
                if not info["any_value"]:
                    (missing_required if level == "Required" else missing_recommended).append(info["name"])

            if missing_required or missing_recommended:
                out_rows.append([
                    sku, wpid_of.get(sku, ""), title_of.get(sku, ""),
                    "; ".join(sorted(set(missing_required))),
                    "; ".join(sorted(set(missing_recommended))),
                ])
                total_missing_required += len(missing_required)
        wb.close()

    out_path = os.path.join(output_dir, f"missing_aspects_{country}.csv")
    with open(out_path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["sku", "wpid", "title", "missing_required", "missing_recommended"])
        w.writerows(out_rows)

    print(f"\ndone: {len(seen_skus)} SKUs checked, {len(out_rows)} with at least one gap, "
          f"{total_missing_required} total missing-required fields -> {out_path}")


if __name__ == "__main__":
    main()
